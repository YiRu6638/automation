import openpyxl
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
data_path = os.path.join(script_dir, "./", "output-record.xlsx")
template_path = os.path.join(script_dir, "template", "Name_Personal_ITC_YA2025_R.xlsx")
template2_path = os.path.join(script_dir, "template", "Name_Personal_ITC_YA2025_NR.xlsx")
data = openpyxl.load_workbook(data_path)
data_sheet = data.active
 
# Ensure output directory exists for generated tax computation files
tax_comps_dir = os.path.join(script_dir, "tax_comps")
os.makedirs(tax_comps_dir, exist_ok=True)



def generate_tax_comp():
    for row in range(3, data_sheet.max_row + 1):
        row_data = list(data_sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=26, values_only=True))[0]

        if row_data[21] == 'R':
            template_workbook = openpyxl.load_workbook(template_path)
            template_sheet = template_workbook["Tax Comp"]
            
            template_sheet['C4'] = row_data[1] # Name
            template_sheet['C5'] = row_data[0] # Tax Identification Number
            template_sheet['E14'] = row_data[3] # Gross salary
            template_sheet['E15'] = row_data[4] # Bonuses/director fees
            template_sheet['E16'] = row_data[5] # Gross tips
            template_sheet['E17'] = row_data[6] 
            template_sheet['E18'] = row_data[7]
            template_sheet['E19'] = row_data[8]
            template_sheet['E20'] = row_data[9]
            template_sheet['E21'] = row_data[10]
            template_sheet['E22'] = row_data[11]
            template_sheet['E23'] = row_data[12]
            template_sheet['E24'] = row_data[13]
            template_sheet['E27'] = row_data[25] * -1 if row_data[25] else 0
            template_sheet['E39'] = row_data[18] * -1 if row_data[18] else 0
            template_sheet['E42'] = row_data[17] * -1 if row_data[17] else 0
            
            name = row_data[1].replace(" ", "")
            
            new_file_name = name + "_Personal_ITC_YA2025.xlsx"
            new_filepath = os.path.join(tax_comps_dir, new_file_name)
            template_workbook.save(new_filepath)
        else:
            template_workbook = openpyxl.load_workbook(template2_path)
            template_sheet = template_workbook.active
            
            template_sheet['C4'] = row_data[1]
            template_sheet['C5'] = row_data[0]
            template_sheet['E14'] = row_data[3]
            template_sheet['E15'] = row_data[4]
            template_sheet['E16'] = row_data[5]
            template_sheet['E17'] = row_data[6]
            template_sheet['E18'] = row_data[7]
            template_sheet['E19'] = row_data[8]
            template_sheet['E20'] = row_data[9]
            template_sheet['E21'] = row_data[10]
            template_sheet['E22'] = row_data[11]
            template_sheet['E23'] = row_data[12]
            template_sheet['E24'] = row_data[13]
            template_sheet['E32'] = row_data[18] * -1 if row_data[18] else 0
            template_sheet['E35'] = row_data[17] * -1 if row_data[17] else 0

            name = row_data[1].replace(" ", "")
            new_file_name = name + "_Personal_ITC_YA2025_Non-resident.xlsx"

            new_filepath = os.path.join(tax_comps_dir, new_file_name)
            template_workbook.save(new_filepath)

if __name__ == "__main__":
    generate_tax_comp()
    pass